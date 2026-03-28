#!/usr/bin/env python3
"""
ClaudeSec — 통합 대시보드 빌드

Google Sheets 3개 + Datadog + Prowler + ClaudeSec 스캔 결과를
수집하여 대시보드 HTML에 데이터를 주입합니다.

사용법:
  source .venv-asset/bin/activate
  python3 scripts/build-dashboard.py
"""

import json
import os
import subprocess
import sys
import urllib.request
import urllib.error
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Any, TypedDict

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "scanner" / "lib"))
from csp_utils import generate_nonce, inject_csp_nonce

import gspread

try:
    import openpyxl
except ImportError:
    openpyxl = None

# ── 설정 ──────────────────────────────────────────────────────────────────

ROOT = Path(__file__).resolve().parent.parent
ASSETS_DIR = ROOT / ".claudesec-assets"
ASSETS_DIR.mkdir(parents=True, exist_ok=True)

sheet_ids: dict[str, str] = {}


class NotionAudit(TypedDict):
    date: str
    title: str
    status: str
    priority: str
    impact: str
    tags: list[str]
    summary: str
    url: str
    task_id: str


KST = timezone(timedelta(hours=9))
NOW = datetime.now(KST).strftime("%Y-%m-%d %H:%M KST")


# .env 로드 (CLAUDESEC_ENV_FILE > 프로젝트 루트 .env > ~/Desktop/.env)
def load_env() -> dict[str, str]:
    env: dict[str, str] = {}
    candidates = [
        Path(os.environ["CLAUDESEC_ENV_FILE"]) if os.environ.get("CLAUDESEC_ENV_FILE") else None,
        ROOT / ".env",
        Path(os.path.expanduser("~/Desktop/.env")),
    ]
    for p in candidates:
        if p and p.exists():
            for line in p.read_text().splitlines():
                if "=" in line and not line.startswith("#"):
                    k, v = line.split("=", 1)
                    env[k.strip()] = v.strip()
            break
    return env


env_vars: dict[str, str] = load_env()
DD_API_KEY = env_vars.get("datadog_key_credential", "")
DD_APP_KEY = env_vars.get("datadog_app_key_credential", "")


# ── Datadog API ───────────────────────────────────────────────────────────


def dd_get(path: str, params: str = "") -> dict[str, object]:
    url = f"https://api.datadoghq.com{path}"
    if params:
        url += f"?{params}"
    req = urllib.request.Request(
        url,
        headers={
            "DD-API-KEY": DD_API_KEY,
            "DD-APPLICATION-KEY": DD_APP_KEY,
            "Content-Type": "application/json",
        },
    )
    try:
        with urllib.request.urlopen(
            req, timeout=30
        ) as r:  # nosemgrep: dynamic-urllib-use-detected — trusted internal API URLs
            payload = json.loads(r.read().decode("utf-8"))
            return payload if isinstance(payload, dict) else {}
    except Exception as e:
        print(f"  DD API 오류 ({path}): {e}")
        return {}


def dd_post(path: str, body: object) -> dict[str, object]:
    url = f"https://api.datadoghq.com{path}"
    req = urllib.request.Request(
        url,
        json.dumps(body).encode(),
        method="POST",
        headers={
            "DD-API-KEY": DD_API_KEY,
            "DD-APPLICATION-KEY": DD_APP_KEY,
            "Content-Type": "application/json",
        },
    )
    try:
        with urllib.request.urlopen(
            req, timeout=30
        ) as r:  # nosemgrep: dynamic-urllib-use-detected — trusted internal API URLs
            payload = json.loads(r.read().decode("utf-8"))
            return payload if isinstance(payload, dict) else {}
    except Exception as e:
        print(f"  DD API 오류 ({path}): {e}")
        return {}


# ── 수집 함수들 ──────────────────────────────────────────────────────────


def collect_datadog():
    """Datadog 호스트 + 보안 시그널 수집"""
    if not DD_API_KEY:
        print("  Datadog API Key 없음, 스킵")
        return [], []

    print("  Datadog 호스트...")
    raw = dd_get("/api/v1/hosts", "count=200")
    hosts = []
    host_list = raw.get("host_list", [])
    if not isinstance(host_list, list):
        host_list = []
    for h in host_list:
        if not isinstance(h, dict):
            continue
        tags = {}
        tags_by_source = h.get("tags_by_source", {})
        if not isinstance(tags_by_source, dict):
            tags_by_source = {}
        for src, tlist in tags_by_source.items():
            if not isinstance(tlist, list):
                continue
            for t in tlist:
                if ":" in t:
                    k, v = t.split(":", 1)
                    tags[k] = v
        hosts.append(
            {
                "name": h.get("name", ""),
                "instance_type": tags.get("instance-type", ""),
                "region": tags.get("region", ""),
                "cluster": tags.get(
                    "aws_eks_cluster-name", tags.get("eks_eks-cluster-name", "")
                ),
                "nodepool": tags.get("karpenter.sh/nodepool", ""),
                "env": tags.get("env", ""),
                "aws_alias": tags.get("aws_alias", ""),
                "aws_account": tags.get("aws_account", ""),
                "up": h.get("up", False),
                "agent_version": h.get("meta", {}).get("agent_version", ""),
            }
        )
    print(f"    {len(hosts)}개")

    print("  Datadog 보안 시그널 (14일)...")
    raw = dd_post(
        "/api/v2/security_monitoring/signals/search",
        {
            "filter": {
                "from": "now-14d",
                "to": "now",
                "query": "status:(high OR critical OR medium)",
            },
            "sort": "timestamp",
            "page": {"limit": 100},
        },
    )
    signals = []
    signal_data = raw.get("data", [])
    if not isinstance(signal_data, list):
        signal_data = []
    for s in signal_data:
        if not isinstance(s, dict):
            continue
        a = s.get("attributes", {})
        if not isinstance(a, dict):
            a = {}
        signals.append(
            {
                "id": s.get("id", "")[:30],
                "title": a.get("message", "")[:200],
                "severity": a.get("severity", ""),
                "status": a.get("status", ""),
                "timestamp": a.get("timestamp", ""),
                "tags": a.get("tags", [])[:5],
            }
        )
    print(f"    {len(signals)}건")
    return hosts, signals


def collect_prowler():
    """Prowler 결과 요약"""
    print("  Prowler...")
    pdir = ROOT / ".claudesec-prowler"
    if not pdir.exists():
        return {}
    findings = []
    for f in pdir.glob("*.ocsf.json"):
        try:
            data = json.loads(f.read_text())
            findings.extend(data if isinstance(data, list) else [data])
        except json.JSONDecodeError:
            # Handle NDJSON or multi-object files (e.g., prowler-iac.ocsf.json)
            decoder = json.JSONDecoder()
            raw = f.read_text().strip()
            pos = 0
            while pos < len(raw):
                try:
                    obj, end = decoder.raw_decode(raw, pos)
                    if isinstance(obj, list):
                        findings.extend(obj)
                    else:
                        findings.append(obj)
                    pos = end
                except json.JSONDecodeError:
                    pos += 1
        except Exception:
            pass

    failures = [f for f in findings if f.get("status_code") == "FAIL"]

    # Use Prowler's severity text as-is (trusted), include Critical+High in samples
    def _is_critical_or_high(f):
        return (f.get("severity") or "").lower() in ("high", "critical")

    # Truncated check names → verified full names on hub.prowler.com
    _PROWLER_NAME_FIX = {
        "core_minimize_containers_added_capabiliti": "core_minimize_containers_added_capabilities",
        "iam_aws_attached_policy_no_administrative_privil": "iam_aws_attached_policy_no_administrative_privileges",
    }
    # Check names known to NOT exist on hub.prowler.com
    _PROWLER_HUB_BLOCKLIST = {"iac-branch", "slack-web-hook"}

    def _prowler_hub_url(check_id: str) -> str:
        """Generate validated Prowler Hub URL from check ID."""
        import re as _re

        if not check_id:
            return ""
        name = _re.sub(r"^prowler-[a-z]+-", "", check_id)
        name = _re.sub(r"-iac-branch-\.[a-z0-9/]+$", "", name)
        name = _re.sub(r"-\d{12}.*$", "", name)
        name = _re.sub(r"-[0-9a-f]{5,}$", "", name)
        # Skip blocklisted patterns
        if any(bl in name for bl in _PROWLER_HUB_BLOCKLIST):
            return ""
        # Fix known truncated names
        name = _PROWLER_NAME_FIX.get(name, name)
        return f"https://hub.prowler.com/check/{name}" if name else ""

    fail_sev = {}
    for f in failures:
        s = f.get("severity", "unknown")
        fail_sev[s] = fail_sev.get(s, 0) + 1

    samples = []
    seen_msgs = set()
    for f in failures:
        if _is_critical_or_high(f):
            msg = f.get("message", f.get("status_detail", ""))[:150]
            if msg in seen_msgs:
                continue
            seen_msgs.add(msg)
            um = f.get("unmapped", {})
            check_id = (
                f.get("finding_info", {}).get("uid", "")
                if isinstance(f.get("finding_info"), dict)
                else ""
            )[:60]
            samples.append(
                {
                    "message": msg,
                    "severity": f.get("severity", "High"),
                    "provider": um.get("provider", "") if isinstance(um, dict) else "",
                    "check": check_id,
                    "hub_url": _prowler_hub_url(check_id),
                }
            )
    # Sort: Critical first, then High
    samples.sort(key=lambda x: (0 if x["severity"] == "Critical" else 1, x["message"]))

    # Group by provider
    by_provider = {}
    for f in failures:
        um = f.get("unmapped", {})
        prov = (
            um.get("provider", "") if isinstance(um, dict) else ""
        ).lower() or "unknown"
        if prov not in by_provider:
            by_provider[prov] = {"total": 0, "fail": 0, "pass": 0, "critical_fails": []}
        by_provider[prov]["fail"] += 1

    # Also count passes per provider
    for f in findings:
        um = f.get("unmapped", {})
        prov = (
            um.get("provider", "") if isinstance(um, dict) else ""
        ).lower() or "unknown"
        if prov not in by_provider:
            by_provider[prov] = {"total": 0, "fail": 0, "pass": 0, "critical_fails": []}
        by_provider[prov]["total"] += 1
        if f.get("status_code") != "FAIL":
            by_provider[prov]["pass"] += 1

    # Add critical fails samples per provider
    for f in failures:
        if _is_critical_or_high(f):
            um = f.get("unmapped", {})
            prov = (
                um.get("provider", "") if isinstance(um, dict) else ""
            ).lower() or "unknown"
            if prov in by_provider and len(by_provider[prov]["critical_fails"]) < 10:
                cid = (
                    f.get("finding_info", {}).get("uid", "")[:60]
                    if isinstance(f.get("finding_info"), dict)
                    else ""
                )
                by_provider[prov]["critical_fails"].append(
                    {
                        "message": f.get("message", f.get("status_detail", ""))[:150],
                        "severity": f.get("severity", "High"),
                        "check": cid,
                        "hub_url": _prowler_hub_url(cid),
                    }
                )

    # Medium + High severity samples (not already in critical_fails_sample)
    medium_samples = []
    critical_msgs = {s["message"] for s in samples}
    for f in failures:
        msg = f.get("message", f.get("status_detail", ""))[:150]
        if (
            f.get("severity") in ("Medium", "High")
            and msg not in critical_msgs
            and len(medium_samples) < 50
        ):
            um = f.get("unmapped", {})
            mid = (
                f.get("finding_info", {}).get("uid", "")
                if isinstance(f.get("finding_info"), dict)
                else ""
            )[:60]
            medium_samples.append(
                {
                    "message": msg,
                    "severity": f.get("severity", "Medium"),
                    "provider": um.get("provider", "") if isinstance(um, dict) else "",
                    "check": mid,
                    "hub_url": _prowler_hub_url(mid),
                }
            )

    # Normalize provider keys: merge kubernetes/k8s, keep iac as "IaC"
    PROVIDER_MERGE = {
        "k8s": "kubernetes",
        "iac": "IaC",
    }
    normalized_by_provider = {}
    for prov, stats in by_provider.items():
        canonical = PROVIDER_MERGE.get(prov, prov)
        if canonical not in normalized_by_provider:
            normalized_by_provider[canonical] = {
                "total": 0,
                "fail": 0,
                "pass": 0,
                "critical_fails": [],
            }
        normalized_by_provider[canonical]["total"] += stats["total"]
        normalized_by_provider[canonical]["fail"] += stats["fail"]
        normalized_by_provider[canonical]["pass"] += stats["pass"]
        normalized_by_provider[canonical]["critical_fails"].extend(
            stats["critical_fails"]
        )
    by_provider = normalized_by_provider

    result = {
        "total": len(findings),
        "pass": len(findings) - len(failures),
        "fail": len(failures),
        "fail_by_severity": fail_sev,
        "critical_fails_sample": samples[:30],
        "medium_fails_sample": medium_samples,
        "by_provider": by_provider,
    }
    print(f"    총 {len(findings)}건 (FAIL: {len(failures)})")
    return result


def get_notion_cache_path() -> Path:
    cache_override = env_vars.get("NOTION_MCP_CACHE_PATH", "") or os.environ.get(
        "NOTION_MCP_CACHE_PATH", ""
    )
    return (
        Path(cache_override).expanduser()
        if cache_override
        else ASSETS_DIR / "notion-security-audits.json"
    )


def load_cached_notion_audits(cache_path: Path) -> list[NotionAudit]:
    if not cache_path.exists():
        return []
    try:
        audits = json.loads(cache_path.read_text())
    except Exception:
        return []

    if not isinstance(audits, list):
        return []

    normalized: list[NotionAudit] = []
    for item in audits:
        if not isinstance(item, dict):
            continue
        normalized.append(
            {
                "date": item.get("date", ""),
                "title": item.get("title", ""),
                "status": item.get("status", "완료"),
                "priority": item.get("priority", ""),
                "impact": item.get("impact", ""),
                "tags": item.get("tags", [])
                if isinstance(item.get("tags", []), list)
                else [],
                "summary": item.get("summary", ""),
                "url": item.get("url", ""),
                "task_id": item.get("task_id", ""),
            }
        )
    return normalized


def merge_notion_cache_fields(
    audits: list[NotionAudit], cache_path: Path
) -> list[NotionAudit]:
    cached = load_cached_notion_audits(cache_path)
    cache_map: dict[str, NotionAudit] = {
        item["url"]: item for item in cached if item["url"]
    }
    merged: list[NotionAudit] = []
    for audit in audits:
        cached_entry = cache_map.get(audit["url"])
        merged.append(
            {
                "date": audit["date"],
                "title": audit["title"],
                "status": audit["status"],
                "priority": audit["priority"]
                or (cached_entry["priority"] if cached_entry else ""),
                "impact": audit["impact"]
                or (cached_entry["impact"] if cached_entry else ""),
                "tags": audit["tags"] or (cached_entry["tags"] if cached_entry else []),
                "summary": audit["summary"]
                or (cached_entry["summary"] if cached_entry else ""),
                "url": audit["url"],
                "task_id": audit["task_id"],
            }
        )
    return merged


ALLOWED_SYNC_SCRIPTS = {
    "sync-notion-audits-mcp.py",
    "scripts/sync-notion-audits-mcp.py",
}


def run_notion_mcp_sync(cache_path: Path) -> bool:
    sync_command = env_vars.get("NOTION_MCP_SYNC_COMMAND", "") or os.environ.get(
        "NOTION_MCP_SYNC_COMMAND", ""
    )
    if not sync_command:
        return False

    # Validate against allowlist to prevent command injection
    cmd_name = Path(sync_command).name
    if cmd_name not in ALLOWED_SYNC_SCRIPTS and sync_command not in ALLOWED_SYNC_SCRIPTS:
        print(f"    Notion MCP sync: 허용되지 않은 명령 '{cmd_name}'")
        print(f"    허용 목록: {', '.join(sorted(ALLOWED_SYNC_SCRIPTS))}")
        return False

    script_path = ROOT / sync_command if not Path(sync_command).is_absolute() else Path(sync_command)
    if not script_path.exists():
        print(f"    Notion MCP sync: 스크립트 없음 '{script_path}'")
        return False

    print("  Notion MCP sync command 실행...")
    cmd_env = os.environ.copy()
    cmd_env.update(env_vars)
    cmd_env["NOTION_MCP_CACHE_PATH"] = str(cache_path)
    result = subprocess.run(
        ["python3", str(script_path)],
        cwd=ROOT,
        env=cmd_env,
        capture_output=True,
        text=True,
        timeout=120,
        check=False,
    )

    if result.returncode != 0:
        message = (result.stderr or result.stdout).strip()
        print(f"    Notion MCP sync 실패: {message[:200]}")
        return False

    print("    Notion MCP sync 완료")
    return True


def fetch_notion_audits_via_api(cache_path: Path) -> list[NotionAudit]:
    notion_key = env_vars.get("NOTION_API_KEY", "") or os.environ.get(
        "NOTION_API_KEY", ""
    )
    if not notion_key:
        return []

    print("  Notion API로 정기점검 이력 수집...")
    try:
        req = urllib.request.Request(
            "https://api.notion.com/v1/search",
            json.dumps(
                {
                    "query": "보안로그 정기점검",
                    "filter": {"property": "object", "value": "page"},
                    "sort": {
                        "direction": "descending",
                        "timestamp": "last_edited_time",
                    },
                    "page_size": 20,
                }
            ).encode(),
            method="POST",
            headers={
                "Authorization": f"Bearer {notion_key}",
                "Content-Type": "application/json",
                "Notion-Version": "2022-06-28",
            },
        )
        with urllib.request.urlopen(
            req, timeout=30
        ) as r:  # nosemgrep: dynamic-urllib-use-detected — trusted internal API URLs
            payload = json.loads(r.read().decode("utf-8"))

        if not isinstance(payload, dict):
            return []

        audits: list[NotionAudit] = []
        results = payload.get("results", [])
        if not isinstance(results, list):
            return []

        for page in results:
            if not isinstance(page, dict):
                continue
            props = page.get("properties", {})
            if not isinstance(props, dict):
                props = {}
            title_parts = props.get("title", props.get("작업 이름", {}))
            title = ""
            if isinstance(title_parts, dict):
                title_values = title_parts.get("title", [])
                if isinstance(title_values, list):
                    for t in title_values:
                        if isinstance(t, dict):
                            plain_text = t.get("plain_text", "")
                            if isinstance(plain_text, str):
                                title += plain_text

            if "정기점검" not in title:
                continue

            page_id = page.get("id", "")
            created_time = page.get("created_time", "")
            last_edited_time = page.get("last_edited_time", "")
            page_url = page.get("url", "")
            if not isinstance(page_id, str):
                page_id = ""
            if not isinstance(created_time, str):
                created_time = ""
            if not isinstance(last_edited_time, str):
                last_edited_time = ""
            if not isinstance(page_url, str):
                page_url = ""
            created = created_time[:10]
            last_edited = last_edited_time[:10]
            url = page_url or f"https://www.notion.so/{page_id.replace('-', '')}"
            audits.append(
                {
                    "date": last_edited or created,
                    "title": title,
                    "status": "완료",
                    "priority": "",
                    "impact": "",
                    "tags": [],
                    "summary": "",
                    "url": url,
                    "task_id": page_id[:8],
                }
            )

        if audits:
            merged = merge_notion_cache_fields(audits, cache_path)
            cache_path.write_text(json.dumps(merged, ensure_ascii=False, indent=2))
            print(f"    Notion API: {len(merged)}건 (캐시 업데이트)")
            return merged
    except Exception as e:
        print(f"    Notion API 실패: {e}, 캐시 사용")

    return []


def collect_notion_audits() -> list[NotionAudit]:
    cache_path = get_notion_cache_path()

    # 1) MCP sync (preferred)
    if run_notion_mcp_sync(cache_path):
        audits = load_cached_notion_audits(cache_path)
        if audits:
            print(f"    Notion MCP cache: {len(audits)}건")
            return audits

    # 2) Live API (before stale cache)
    audits = fetch_notion_audits_via_api(cache_path)
    if audits:
        return audits

    # 3) Stale cache fallback
    audits = load_cached_notion_audits(cache_path)
    if audits:
        print(f"    Notion 정기점검: {len(audits)}건 (캐시, stale)")
        return audits

    return []


def collect_jamf_pcs():
    """Collect Jamf Pro managed PC data from Datadog logs API."""
    if not DD_API_KEY:
        print("  Jamf PC: Datadog API Key 없음, 스킵")
        return []

    print("  Jamf Pro PC 데이터...")
    # Search Datadog logs for Jamf ComputerPolicyFinished events (webhook)
    raw = dd_post(
        "/api/v2/logs/events/search",
        {
            "filter": {
                "from": "now-7d",
                "to": "now",
                "query": "source:jamf*",
            },
            "sort": "-timestamp",
            "page": {"limit": 200},
        },
    )

    pcs = []
    seen = set()
    log_data = raw.get("data", [])
    if not isinstance(log_data, list):
        log_data = []
    for log in log_data:
        if not isinstance(log, dict):
            continue
        attrs = log.get("attributes", {}).get("attributes", {})
        # Jamf webhook structure: event.computer contains device info
        computer = attrs.get("event", {}).get("computer", {})
        if not isinstance(computer, dict):
            computer = attrs  # fallback to flat structure
        if not isinstance(computer, dict):
            continue
        hostname = computer.get(
            "deviceName", computer.get("computer_name", attrs.get("hostname", ""))
        )
        serial = computer.get("serialNumber", computer.get("serial_number", ""))
        key = serial or hostname
        if not key or key in seen:
            continue
        seen.add(key)
        ts = log.get("attributes", {}).get("timestamp", "")
        # Parse user from deviceName pattern "Lv-i <username> <serial>"
        raw_user = computer.get("realName", "") or attrs.get("usr", {}).get("name", "")
        if not raw_user and hostname:
            parts = hostname.split()
            if len(parts) >= 2 and parts[0].lower().startswith("lv"):
                raw_user = parts[1]
        pcs.append(
            {
                "name": hostname,
                "serial": serial,
                "os": computer.get("osVersion", computer.get("os_version", "")),
                "model": computer.get("model", ""),
                "user": raw_user,
                "department": computer.get("department", ""),
                "managed": True,
                "last_checkin": (ts[:19] if ts else ""),
                "source": "jamf",
            }
        )

    # Try full inventory file (from Jamf Pro CSV export)
    inventory_path = ASSETS_DIR / "jamf-full-inventory.json"
    if not pcs and inventory_path.exists():
        try:
            inv = json.loads(inventory_path.read_text())
            pcs = [item for item in inv if item.get("type") == "computer"]
            print(f"    Jamf PC: {len(pcs)}대 (CSV 인벤토리)")
            return pcs
        except Exception:
            pass

    # If no Datadog logs, try loading from cache
    cache_path = ASSETS_DIR / "jamf-computers.json"
    if not pcs and cache_path.exists():
        try:
            pcs = json.loads(cache_path.read_text())
            print(f"    Jamf PC: {len(pcs)}대 (캐시)")
            return pcs
        except Exception:
            pass

    if pcs:
        cache_path.write_text(json.dumps(pcs, ensure_ascii=False, indent=2))

    print(f"    Jamf PC: {len(pcs)}대")
    return pcs


def collect_intune_pcs():
    """Collect Intune managed PC data from cache/CSV."""
    cache_path = ASSETS_DIR / "intune-computers.json"
    if cache_path.exists():
        try:
            pcs = json.loads(cache_path.read_text())
            print(f"  Intune PC: {len(pcs)}대")
            return pcs
        except Exception:
            pass
    print("  Intune PC: 데이터 없음")
    return []


def collect_policies():
    """Load policy/regulation data from cache."""
    cache_path = ASSETS_DIR / "policies.json"
    if cache_path.exists():
        try:
            policies = json.loads(cache_path.read_text())
            total_articles = sum(p.get("total_articles", 0) for p in policies)
            print(f"  규정/지침: {len(policies)}개 ({total_articles}개 조항)")
            return policies
        except Exception:
            pass
    print("  규정/지침: 데이터 없음")
    return []


def collect_sheets():
    """Google Sheets 데이터 수집"""
    print("  Google Sheets 연결...")
    gc = gspread.oauth(
        scopes=[
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive",
        ]
    )

    # 자산관리대장
    sp1 = gc.open_by_key(sheet_ids["자산관리대장"])
    print(f"    '{sp1.title}' 연결")

    # SaaS 파싱 — col[1]=No, col[2]=구분, col[3]=자산명, col[4]=설명,
    #   col[5]=제공사, col[6]=관리주체, col[7]=접근관리방식, col[8]=비고
    ws_saas = sp1.worksheet("8.SaaS")
    raw = ws_saas.get_all_values()
    saas = []
    for row in raw[4:]:  # 데이터는 4번째 행부터 (index 4)
        cells = [c.strip() for c in row]
        if not any(cells):
            continue
        auth_val = cells[7] if len(cells) > 7 else ""
        auth_lower = auth_val.lower()
        sso_linked = auth_lower in ("okta", "okta sso", "google")
        saas.append(
            {
                "no": cells[1] if len(cells) > 1 else "",
                "category": cells[2] if len(cells) > 2 else "",
                "name": cells[3] if len(cells) > 3 else "",
                "description": cells[4] if len(cells) > 4 else "",
                "provider": cells[5] if len(cells) > 5 else "",
                "owner": cells[6] if len(cells) > 6 else "",
                "auth": auth_val,
                "sso": sso_linked,
                "contract_period": cells[9] if len(cells) > 9 else "",
                "note": cells[8] if len(cells) > 8 else "",
            }
        )
    print(f"    SaaS: {len(saas)}개")

    # 라이선스 현황 파싱 (row 0 = 헤더, row 1+ = 데이터)
    ws_lic = sp1.worksheet("라이선스_현황")
    raw = ws_lic.get_all_values()
    license_list = []
    for row in raw[1:]:
        cells = [c.strip() for c in row]
        if not any(cells):
            continue
        license_list.append(
            {
                "name": cells[0] if len(cells) > 0 else "",
                "description": cells[1] if len(cells) > 1 else "",
                "status": cells[2] if len(cells) > 2 else "",
                "admin": cells[3] if len(cells) > 3 else "",
                "department": cells[4] if len(cells) > 4 else "",
                "accounts": cells[5] if len(cells) > 5 else "",
                "active_accounts": cells[6] if len(cells) > 6 else "",
                "users": cells[7] if len(cells) > 7 else "",
                "cycle": cells[8] if len(cells) > 8 else "",
                "cost": cells[9] if len(cells) > 9 else "",
                "contract_period": cells[10] if len(cells) > 10 else "",
            }
        )
    print(f"    라이선스: {len(license_list)}개")

    # 자산 분류별 수량
    asset_counts = {}
    sheet_map = {
        "1.서버": "서버",
        "2.정보보호시스템": "정보보호시스템",
        "3.네트워크 장비": "네트워크 장비",
        "4. DBMS": "DBMS",
        "5.PC": "PC",
        "6.홈페이지": "홈페이지",
        "7. 개인정보": "개인정보",
        "8.SaaS": "SaaS",
    }
    for ws_item in sp1.worksheets():
        if ws_item.title in sheet_map:
            vals = ws_item.get_all_values()
            data_rows = (
                len([r for r in vals[8:] if any(c.strip() for c in r)])
                if len(vals) > 8
                else 0
            )
            asset_counts[sheet_map[ws_item.title]] = data_rows
    asset_counts["라이선스"] = len(license_list)
    print(f"    자산 분류: {asset_counts}")

    # ── 인프라 상세 파싱 ──
    # 서버 (EC2) — row 7=헤더(col1:구분,col2:자산코드,...), row 8+=데이터
    raw_srv = sp1.worksheet("1.서버").get_all_values()
    servers = []
    for row in raw_srv[8:]:
        c = [x.strip() for x in row]
        if not c[1] and not c[2]:
            continue
        servers.append(
            {
                "os_type": c[1],
                "instance_id": c[2],
                "name": c[3],
                "account": c[4],
                "os": c[5],
                "instance_type": c[6],
                "count": c[7],
                "region": c[8],
            }
        )
    print(f"    서버 (EC2): {len(servers)}개")

    # DBMS — row 7=헤더, row 8+=데이터
    raw_db = sp1.worksheet("4. DBMS").get_all_values()
    databases = []
    for row in raw_db[8:]:
        c = [x.strip() for x in row]
        if not c[1] and not c[2]:
            continue
        databases.append(
            {
                "type": c[1],
                "name": c[2],
                "version": c[3],
                "engine": c[4],
                "count": c[6],
                "platform": c[7],
                "region": c[8],
                "account": c[10] if len(c) > 10 else "",
            }
        )
    print(f"    DBMS: {len(databases)}개")

    # 정보보호시스템 — row 7=헤더, row 8+=데이터
    raw_sec = sp1.worksheet("2.정보보호시스템").get_all_values()
    sec_systems = []
    for row in raw_sec[8:]:
        c = [x.strip() for x in row]
        if not c[1] and not c[2]:
            continue
        sec_systems.append(
            {
                "category": c[1],
                "name": c[2],
                "description": c[3],
                "type": c[4],
                "provider": c[5],
                "domain": c[6],
                "owner": c[7],
                "auth": c[8],
            }
        )
    print(f"    정보보호시스템: {len(sec_systems)}개")

    # 네트워크 장비 — row 7=헤더, row 8+=데이터
    raw_net = sp1.worksheet("3.네트워크 장비").get_all_values()
    networks = []
    for row in raw_net[8:]:
        c = [x.strip() for x in row]
        if not c[1] and not c[2]:
            continue
        networks.append(
            {
                "type": c[1],
                "resource_id": c[2],
                "name": c[3],
                "description": c[4],
                "region": c[5],
                "status": c[7],
                "account": c[9] if len(c) > 9 else "",
            }
        )
    print(f"    네트워크 장비: {len(networks)}개")

    # AI 구독 현황
    sp2 = gc.open_by_key(sheet_ids["AI구독현황"])
    print(f"    '{sp2.title}' 연결")
    ws_ai = sp2.worksheets()[0]
    raw_ai = ws_ai.get_all_values()
    ai_list = []
    for row in raw_ai[3:]:  # 데이터는 3번째 행부터
        cells = [c.strip() for c in row]
        if not any(cells):
            continue
        ai_list.append(
            {
                "name": cells[0] if len(cells) > 0 else "",
                "description": cells[1] if len(cells) > 1 else "",
                "quantity": cells[2] if len(cells) > 2 else "",
                "department": cells[3] if len(cells) > 3 else "",
            }
        )
    print(f"    AI 구독: {len(ai_list)}개")

    # SaaS 비용 (.xlsx via Drive API download)
    cost_data: dict[str, Any] = {"summary": [], "details": {}}
    xlsx_id = env_vars.get("COST_XLSX_ID", "")
    if openpyxl:
        try:
            print("  SaaS 비용 시트 다운로드 중...")
            resp = gc.http_client.request(
                "get",
                f"https://www.googleapis.com/drive/v3/files/{xlsx_id}?alt=media&supportsAllDrives=true",
            )
            tmp_path = Path("/tmp/claudesec-cost.xlsx")
            tmp_path.write_bytes(resp.content)

            wb = openpyxl.load_workbook(tmp_path, data_only=True)

            # 요약 시트 — 3개 섹션 분리 (소프트웨어별/사용자별/부서별)
            ws_sum = wb["요약"]
            section = "software"
            cost_data["by_user"] = []
            cost_data["by_department"] = []
            for row in ws_sum.iter_rows(min_row=3, values_only=True):
                cells = list(row)
                label = str(cells[0] or "").strip()
                if not label:
                    continue
                if label in ("사용자별 요약", "사용자"):
                    section = "user"
                    continue
                if label in ("부서별 요약", "부서"):
                    section = "department"
                    continue
                if label in ("합계", "소프트웨어"):
                    continue
                vals = {
                    "jan": cells[1] if isinstance(cells[1], (int, float)) else 0,
                    "feb": cells[2] if isinstance(cells[2], (int, float)) else 0,
                    "mar": cells[3] if isinstance(cells[3], (int, float)) else 0,
                    "total": cells[4] if isinstance(cells[4], (int, float)) else 0,
                }
                if section == "software":
                    cost_data["summary"].append({"software": label, **vals})
                elif section == "user":
                    cost_data["by_user"].append({"user": label, **vals})
                elif section == "department":
                    cost_data["by_department"].append({"department": label, **vals})

            # 월별 상세
            for month_name in ["2026-01", "2026-02", "2026-03"]:
                if month_name not in wb.sheetnames:
                    continue
                ws_m = wb[month_name]
                rows_m = []
                for row in ws_m.iter_rows(min_row=5, values_only=True):
                    cells = list(row)
                    if not cells[0]:
                        continue
                    # 합계 행 스킵
                    sw = str(cells[0]).strip()
                    if sw in ("합계",):
                        continue
                    rows_m.append(
                        {
                            "software": sw,
                            "user": str(cells[1] or ""),
                            "shared": str(cells[2] or ""),
                            "department": str(cells[3] or ""),
                            "amount": cells[4]
                            if isinstance(cells[4], (int, float))
                            else 0,
                            "date": str(cells[5] or ""),
                            "vendor": str(cells[6] or ""),
                            "memo": str(cells[7] or ""),
                        }
                    )
                cost_data["details"][month_name] = rows_m

            # ── 중복/유사 항목 통합 정리 ──
            # 1) summary: 유사 소프트웨어명 통합
            SW_MERGE = {
                "GWS(GoogleWorkSpace)": "Google Workspace",
                "GitHub (Copilot/Actions)": "GitHub",
                "Microsoft Office 365": "Microsoft 365 / Intune",
            }
            merged_summary: dict[str, dict[str, object]] = {}
            for s in cost_data["summary"]:
                software_name = str(s.get("software", ""))
                name = SW_MERGE.get(software_name, software_name)
                if name in merged_summary:
                    for k in ("jan", "feb", "mar", "total"):
                        merged_summary[name][k] = merged_summary[name].get(
                            k, 0
                        ) + s.get(k, 0)
                else:
                    merged_summary[name] = {**s, "software": name}
            cost_data["summary"] = list(merged_summary.values())

            # 2) details: 동일 월+소프트웨어+사용자+금액 완전 중복 제거
            for month_key in list(cost_data["details"].keys()):
                rows = cost_data["details"][month_key]
                seen = set()
                deduped = []
                for r in rows:
                    # 소프트웨어명 통합
                    r["software"] = SW_MERGE.get(r["software"], r["software"])
                    key = (r["software"], r.get("user", ""), r.get("amount", 0))
                    if key in seen:
                        continue
                    seen.add(key)
                    deduped.append(r)
                cost_data["details"][month_key] = deduped

            # 3) Apply proxy-payment overrides from config (if present)
            proxy_overrides_raw = json.loads(
                os.environ.get("COST_PROXY_OVERRIDES", "[]")
            )
            proxy_overrides = (
                proxy_overrides_raw if isinstance(proxy_overrides_raw, list) else []
            )
            for month_key, rows in cost_data["details"].items():
                for r in rows:
                    for ovr in proxy_overrides:
                        if r["software"] == ovr.get("software") and r.get(
                            "user"
                        ) == ovr.get("user"):
                            r["memo"] = (
                                r.get("memo", "") + " " + ovr.get("memo", "")
                            ).strip()
                            if ovr.get("department"):
                                r["department"] = ovr["department"]

            print(
                f"    SaaS 비용: {len(cost_data['summary'])}종, 월별 상세 {sum(len(v) for v in cost_data['details'].values())}건 (중복 정리 완료)"
            )
        except Exception as e:
            print(f"    SaaS 비용 수집 실패: {e}")
    else:
        print("  openpyxl 미설치, SaaS 비용 스킵")

    infra = {
        "servers": servers,
        "databases": databases,
        "sec_systems": sec_systems,
        "networks": networks,
    }

    # SentinelOne 에이전트 + 위협 + 교차 검증 (캐시)
    s1_path = ASSETS_DIR / "sentinelone-agents.json"
    s1_agents: list[dict[str, object]] = []
    if s1_path.exists():
        loaded_agents = json.loads(s1_path.read_text())
        if isinstance(loaded_agents, list):
            s1_agents = [item for item in loaded_agents if isinstance(item, dict)]
        print(f"    SentinelOne: {len(s1_agents)}대")

    s1_threats_path = ASSETS_DIR / "sentinelone-threats.json"
    s1_threats: list[dict[str, object]] = []
    if s1_threats_path.exists():
        loaded_threats = json.loads(s1_threats_path.read_text())
        if isinstance(loaded_threats, list):
            s1_threats = [item for item in loaded_threats if isinstance(item, dict)]

    ep_xv_path = ASSETS_DIR / "endpoint-crossverify.json"
    ep_crossverify: dict[str, object] = {}
    if ep_xv_path.exists():
        loaded_crossverify = json.loads(ep_xv_path.read_text())
        if isinstance(loaded_crossverify, dict):
            ep_crossverify = loaded_crossverify

    return (
        saas,
        license_list,
        ai_list,
        asset_counts,
        cost_data,
        infra,
        s1_agents,
        s1_threats,
        ep_crossverify,
    )


def _is_karpenter_node(name: str) -> bool:
    """Karpenter 동적 노드 여부 판별 (ip-10-* 패턴)"""
    import re

    if not name:
        return False
    return bool(re.match(r"ip-10-\d+-\d+-\d+\.", name))


def _cross_verify_ec2(
    aws_ec2: list[dict[str, object]], sheet_servers: list[dict[str, object]]
) -> dict[str, Any]:
    """AWS EC2 실시간 vs 자산관리대장 서버 교차 검증 (Karpenter 노드 분리)"""
    # Karpenter 노드와 고정 서버 분리
    aws_static: dict[str, dict[str, object]] = {}
    aws_karpenter: list[dict[str, object]] = []
    for e in aws_ec2:
        iid = e.get("InstanceId", "")
        name = e.get("Name", "")
        if not isinstance(iid, str) or not iid:
            continue
        if not isinstance(name, str):
            name = ""
        if _is_karpenter_node(name):
            aws_karpenter.append(
                {
                    "id": iid,
                    "name": name,
                    "type": e.get("Type", ""),
                    "profile": e.get("_profile", ""),
                }
            )
        else:
            aws_static[iid] = e

    # 시트에서도 Karpenter/고정 분리
    sheet_static: dict[str, dict[str, object]] = {}
    sheet_karpenter: list[dict[str, object]] = []
    for s in sheet_servers:
        iid = s.get("instance_id", "")
        name = s.get("name", "")
        if not isinstance(iid, str) or not iid:
            continue
        if not isinstance(name, str):
            name = ""
        if _is_karpenter_node(name):
            sheet_karpenter.append(
                {
                    "id": iid,
                    "name": name,
                    "type": s.get("instance_type", ""),
                    "account": s.get("account", ""),
                }
            )
        else:
            sheet_static[iid] = s

    # 고정 서버 교차 검증
    aws_only: list[dict[str, object]] = []
    for iid, e in aws_static.items():
        if iid not in sheet_static:
            aws_only.append(
                {
                    "id": iid,
                    "name": e.get("Name", ""),
                    "type": e.get("Type", ""),
                    "profile": e.get("_profile", ""),
                    "status": "미등록",
                }
            )

    sheet_only: list[dict[str, object]] = []
    for iid, s in sheet_static.items():
        if iid not in aws_static:
            sheet_only.append(
                {
                    "id": iid,
                    "name": s.get("name", ""),
                    "type": s.get("instance_type", ""),
                    "account": s.get("account", ""),
                    "status": "종료/교체",
                }
            )

    matched = len(set(aws_static) & set(sheet_static))

    result = {
        "aws_total": len(aws_ec2),
        "aws_static": len(aws_static),
        "aws_karpenter": len(aws_karpenter),
        "sheet_total": len(sheet_servers),
        "sheet_static": len(sheet_static),
        "sheet_karpenter": len(sheet_karpenter),
        "matched": matched,
        "aws_only": aws_only,
        "sheet_only": sheet_only,
        "aws_only_count": len(aws_only),
        "sheet_only_count": len(sheet_only),
        "karpenter_nodes": aws_karpenter,
        "karpenter_count": len(aws_karpenter),
    }
    print(
        f"  교차검증: 고정 서버 AWS {len(aws_static)} vs 시트 {len(sheet_static)}, "
        f"일치 {matched}, 미등록 {len(aws_only)}, 종료 {len(sheet_only)}"
    )
    print(f"  Karpenter: AWS {len(aws_karpenter)}개 동적 노드 (교차검증 제외)")
    return result


def load_aws_live_data():
    """Load AWS describe results from .claudesec-assets/aws-*.json files"""
    result = {"ec2": [], "rds": [], "elasticache": [], "s3": [], "eks": []}
    aws_profiles = [
        p.strip() for p in os.environ.get("AWS_PROFILES", "").split(",") if p.strip()
    ]
    if not aws_profiles:
        # Fallback: scan .claudesec-assets for aws-ec2-*.json files
        for f in sorted(ASSETS_DIR.glob("aws-ec2-*.json")):
            p = f.stem.replace("aws-ec2-", "")
            if p not in aws_profiles:
                aws_profiles.append(p)
    for profile in aws_profiles:
        for rtype in ["ec2", "rds", "rds-clusters", "elasticache", "s3", "eks"]:
            fpath = ASSETS_DIR / f"aws-{rtype}-{profile}.json"
            if not fpath.exists():
                continue
            try:
                data = json.loads(fpath.read_text())
                if isinstance(data, list):
                    for item in data:
                        item["_profile"] = profile
                    key = rtype.replace("-clusters", "")
                    if key not in result:
                        result[key] = []
                    result[key].extend(data)
                elif isinstance(data, dict) and "clusters" in data:
                    for c in data["clusters"]:
                        result["eks"].append({"name": c, "_profile": profile})
            except Exception:
                pass
    for k, v in result.items():
        if v:
            print(f"    AWS {k}: {len(v)}개")
    return result


def collect_scan_history():
    """Collect scan history from .claudesec-history/"""
    hist_dir = ROOT / ".claudesec-history"
    if not hist_dir.exists():
        return []
    history = []
    for f in sorted(hist_dir.glob("scan-*.json")):
        try:
            data = json.loads(f.read_text())
            history.append(data)
        except Exception:
            pass
    # Sort by timestamp, keep last 30
    history.sort(key=lambda x: x.get("timestamp", ""))
    return history[-30:]


# ── 메인 ──────────────────────────────────────────────────────────────────


def main():
    print("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
    print("  ClaudeSec 통합 대시보드 빌드")
    print(f"  {NOW}")
    print("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
    print()
    # Load sheet IDs from env
    global sheet_ids
    sheet_ids = {
        "자산관리대장": env_vars.get("ASSET_SHEET_ID", ""),
        "AI구독현황": env_vars.get("AI_SHEET_ID", ""),
    }
    if not sheet_ids["자산관리대장"]:
        print("  ⚠ ASSET_SHEET_ID 환경변수 없음 — ~/Desktop/.env에 설정 필요")

    print("▶ 데이터 수집")

    # 수집
    dd_hosts, dd_signals = collect_datadog()
    prowler = collect_prowler()
    notion_audits = collect_notion_audits()
    jamf_pcs = collect_jamf_pcs()
    intune_pcs = collect_intune_pcs()
    policies = collect_policies()
    # Load mobile devices if available
    jamf_mobiles = []
    mobile_inv = ASSETS_DIR / "jamf-full-inventory.json"
    if mobile_inv.exists():
        try:
            all_inv = json.loads(mobile_inv.read_text())
            jamf_mobiles = [item for item in all_inv if item.get("type") == "mobile"]
        except Exception:
            pass
    (
        saas,
        licenses,
        ai_subs,
        asset_counts,
        saas_cost,
        infra,
        s1_agents,
        s1_threats,
        ep_crossverify,
    ) = collect_sheets()

    # ClaudeSec 스캔
    scan = {}
    scan_path = ROOT / "scan-report.json"
    if scan_path.exists():
        scan = json.loads(scan_path.read_text())
        # Normalize: scanner outputs "results", dashboard expects "findings"
        if "results" in scan and "findings" not in scan:
            scan["findings"] = scan.pop("results")
        # Flatten summary fields to top level for dashboard compatibility
        summary = scan.get("summary", {})
        for k in ("score", "grade", "passed", "failed", "warnings", "skipped"):
            if k not in scan and k in summary:
                scan[k] = summary[k]
        # Add best practices reference URLs (override generic OWASP links)
        _CS_REF = {
            "TRIVY-CRIT": "https://aquasecurity.github.io/trivy/latest/docs/scanner/vulnerability/",
            "TRIVY-HIGH": "https://aquasecurity.github.io/trivy/latest/docs/scanner/vulnerability/",
            "TRIVY-MED": "https://aquasecurity.github.io/trivy/latest/docs/scanner/vulnerability/",
            "CICD-005": "https://owasp.org/www-project-devsecops-guideline/latest/02a-Static-Application-Security-Testing",
            "MAC-005": "https://support.apple.com/guide/mac-help/keep-your-mac-up-to-date-mchlpx1065/mac",
            "MAC-006": "https://support.apple.com/guide/mac-help/require-a-password-after-waking-your-mac-mchlp2270/mac",
            "MAC-007": "https://support.apple.com/guide/mac-help/allow-a-remote-computer-to-access-your-mac-mchlp1066/mac",
            "CIS-001": "https://www.cisecurity.org/benchmark/apple_os",
            "CIS-002": "https://www.cisecurity.org/benchmark/apple_os",
            "CIS-003": "https://www.cisecurity.org/benchmark/apple_os",
            "CIS-004": "https://www.cisecurity.org/benchmark/apple_os",
            "CIS-009": "https://www.cisecurity.org/benchmark/apple_os",
            "AI-005": "https://owasp.org/www-project-top-10-for-large-language-model-applications/",
            "AI-008": "https://owasp.org/www-project-top-10-for-large-language-model-applications/",
            "NET-001": "https://cheatsheetseries.owasp.org/cheatsheets/Transport_Layer_Security_Cheat_Sheet.html",
            "CLOUD-010": "https://cloud.google.com/logging/docs/audit",
            "IAM-002": "https://docs.github.com/en/get-started/getting-started-with-git/ignoring-files",
            "SECRETS-004": "https://cheatsheetseries.owasp.org/cheatsheets/Secrets_Management_Cheat_Sheet.html",
            "CODE-SEC-008": "https://cheatsheetseries.owasp.org/cheatsheets/Race_Condition_Cheat_Sheet.html",
            "SAAS-001": "https://docs.github.com/en/code-security/getting-started/github-security-features",
            "SAAS-002": "https://docs.github.com/en/actions/security-for-github-actions/security-guides/security-hardening-for-github-actions",
            "SAAS-API-001": "https://docs.github.com/en/repositories/managing-your-repositorys-settings-and-features/enabling-features-for-your-repository/managing-github-actions-settings-for-a-repository",
            "SAAS-API-002": "https://docs.github.com/en/actions/managing-workflow-runs-and-deployments/managing-workflow-runs",
            "SAAS-ZIA-002": "https://help.zscaler.com/zia/user-management",
            "SAAS-ZIA-003": "https://help.zscaler.com/zia/security-policy-settings",
            "SAAS-ZIA-004": "https://help.zscaler.com/zia/api-key-management",
            "SAAS-ZIA-006": "https://help.zscaler.com/zia/nss-deployment-guide",
            "SAAS-ZIA-007": "https://help.zscaler.com/zia/configuring-saml",
            "PROWLER-GCP-000": "https://docs.prowler.com/projects/prowler-open-source/en/latest/tutorials/gcp/authentication/",
        }
        for f in scan.get("findings", []):
            fid = f.get("id", "")
            if fid in _CS_REF:
                f["ref_url"] = _CS_REF[fid]
        print(f"  ClaudeSec: {scan.get('grade')}/{scan.get('score')}")

    scan_history = collect_scan_history()

    # AWS live 데이터 + 파일 수정 시각
    aws_live = load_aws_live_data()

    # 소스별 수집 타임스탬프
    def file_mtime_str(p):
        """파일 수정 시각을 ISO 문자열로 반환"""
        fp = Path(p)
        if fp.exists():
            return datetime.fromtimestamp(fp.stat().st_mtime, tz=KST).strftime(
                "%Y-%m-%d %H:%M KST"
            )
        return None

    timestamps = {
        "datadog": NOW,  # API 실시간
        "prowler": file_mtime_str(ROOT / ".claudesec-prowler"),
        "claudesec_scan": file_mtime_str(ROOT / "scan-report.json"),
        "google_sheets": NOW,  # API 실시간
        "saas_cost_xlsx": file_mtime_str("/tmp/claudesec-cost.xlsx"),
        "notion": file_mtime_str(ASSETS_DIR / "notion-security-audits.json"),
        "aws_ec2": file_mtime_str(ASSETS_DIR / "aws-ec2-prod.json")
        or file_mtime_str(next(ASSETS_DIR.glob("aws-ec2-*.json"), Path("/dev/null"))),
        "aws_rds": file_mtime_str(ASSETS_DIR / "aws-rds-prod.json")
        or file_mtime_str(next(ASSETS_DIR.glob("aws-rds-*.json"), Path("/dev/null"))),
    }

    # EC2 교차 검증: AWS live vs 자산관리대장
    cross_verify = _cross_verify_ec2(aws_live.get("ec2", []), infra.get("servers", []))

    # 대시보드 데이터 조합
    dashboard_data = {
        "generated_at": NOW,
        "timestamps": timestamps,
        "datadog": {"hosts": dd_hosts, "signals": dd_signals},
        "prowler": prowler,
        "claudesec": scan,
        "scan_history": scan_history,
        "saas": saas,
        "license": licenses,
        "ai": ai_subs,
        "asset_counts": asset_counts,
        "saas_cost": saas_cost,
        "infra": infra,
        "aws_live": aws_live,
        "cross_verify": cross_verify,
        "sentinelone": s1_agents,
        "s1_threats": s1_threats,
        "endpoint_crossverify": ep_crossverify,
        "notion_audits": notion_audits,
        "jamf_pcs": jamf_pcs,
        "intune_pcs": intune_pcs,
        "policies": policies,
        "jamf_mobiles": jamf_mobiles,
    }

    # JSON 저장
    json_path = ASSETS_DIR / "dashboard-data.json"
    json_path.write_text(json.dumps(dashboard_data, ensure_ascii=False, indent=2))
    print(f"\n▶ 데이터 저장: {json_path}")

    # HTML에 데이터 주입
    tmpl_path = ROOT / "claudesec-asset-dashboard.html"
    html = tmpl_path.read_text()

    json_str = json.dumps(dashboard_data, ensure_ascii=False, default=str)

    # 템플릿 내 placeholder 교체
    html = html.replace("__DASHBOARD_DATA__", json_str)
    html = html.replace("const D=/*__DATA__*/null;", f"const D={json_str};")

    # CSP nonce 주입 (빌드마다 새로운 랜덤 nonce 생성)
    nonce = generate_nonce()
    html = inject_csp_nonce(html, nonce)

    out_path = ROOT / "claudesec-asset-dashboard-live.html"
    out_path.write_text(html)
    print(f"▶ 대시보드 생성: {out_path}")

    # 요약
    print()
    print("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
    print("  빌드 완료!")
    print("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
    print(f"  Datadog:  {len(dd_hosts)} 호스트, {len(dd_signals)} 시그널")
    print(f"  Prowler:  {prowler.get('total', 0)} 총, {prowler.get('fail', 0)} FAIL")
    print(f"  ClaudeSec: {scan.get('grade', 'N/A')} ({scan.get('score', 'N/A')}점)")
    print(f"  SaaS:     {len(saas)}개")
    print(f"  라이선스: {len(licenses)}개")
    print(f"  AI 구독:  {len(ai_subs)}개")
    print(f"  SaaS비용: {len(saas_cost.get('summary', []))}종")
    print(f"  스캔이력: {len(scan_history)}건")
    print(f"\n  open {out_path}")


if __name__ == "__main__":
    main()
