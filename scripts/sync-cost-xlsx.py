#!/usr/bin/env python3
"""
ClaudeSec — 통합 대시보드 데이터 동기화

모든 데이터 소스에서 수집하여 dashboard-data.json을 생성합니다:
  1. 자산관리대장 (Google Sheets) → SaaS, 라이선스, 인프라, 자산 분류
  2. AI 구독 현황 (Google Sheets) → AI 구독 목록
  3. 소프트웨어_라이선스_현황.xlsx (로컬 파일) → SaaS 비용 분석
  4. Datadog API → 호스트 + 보안 시그널
  5. AWS CLI → EC2, RDS, EKS describe
  6. Zscaler ZIA API → 보안 포스처
  7. Prowler → 클라우드 보안 스캔 결과
  8. ClaudeSec 스캔 히스토리

사용법:
  python3 scripts/sync-cost-xlsx.py                    # 전체 동기화
  python3 scripts/sync-cost-xlsx.py --cost-only        # 비용 데이터만
  python3 scripts/sync-cost-xlsx.py --sheets-only      # Sheets 데이터만
  python3 scripts/sync-cost-xlsx.py --inject-html      # HTML 대시보드에 데이터 주입
  python3 scripts/sync-cost-xlsx.py --skip-aws         # AWS 수집 스킵
  python3 scripts/sync-cost-xlsx.py --skip-datadog     # Datadog 수집 스킵
  python3 scripts/sync-cost-xlsx.py --skip-zscaler     # Zscaler 수집 스킵

환경변수:
  CLAUDESEC_ENV_FILE — .env 파일 경로 (기본: 프로젝트 .env > ~/Desktop/.env)
  COST_XLSX_PATH     — xlsx 파일 경로 (기본: Google Drive 동기화 경로)
  AWS_PROFILES       — AWS 프로파일 목록 (쉼표 구분)
"""

import json
import os
import re
import subprocess
import sys
import urllib.request
import urllib.error
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError:
    openpyxl = None

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "scanner" / "lib"))
from csp_utils import generate_nonce, inject_csp_nonce

ROOT = Path(__file__).resolve().parent.parent
ASSETS_DIR = ROOT / ".claudesec-assets"
ASSETS_DIR.mkdir(parents=True, exist_ok=True)

KST = timezone(timedelta(hours=9))
NOW = datetime.now(KST).strftime("%Y-%m-%d %H:%M KST")

# Google Sheets IDs (반드시 .env에서 로드 — 코드에 하드코딩 금지)
SHEET_IDS: dict[str, str] = {}  # load_env() 이후 설정됨

# 기본 비용 xlsx 경로 (환경변수로 설정)
DEFAULT_XLSX = Path(os.environ.get("COST_XLSX_PATH", str(Path.home() / "Downloads" / "cost-report.xlsx")))

# 유사 소프트웨어명 통합 매핑
SW_MERGE = {
    "GWS(GoogleWorkSpace)": "Google Workspace",
    "GitHub (Copilot/Actions)": "GitHub",
    "Microsoft Office 365": "Microsoft 365 / Intune",
}


# ═══════════════════════════════════════════════════════════════════════════
# 환경변수 로드
# ═══════════════════════════════════════════════════════════════════════════

def load_env() -> dict[str, str]:
    """CLAUDESEC_ENV_FILE > 프로젝트 루트 .env > ~/Desktop/.env 순서로 로드"""
    env: dict[str, str] = {}
    candidates = [
        Path(os.environ["CLAUDESEC_ENV_FILE"]) if os.environ.get("CLAUDESEC_ENV_FILE") else None,
        ROOT / ".env",
        Path(os.path.expanduser("~/Desktop/.env")),
    ]
    for p in candidates:
        if p and p.exists():
            for line in p.read_text().splitlines():
                if "=" in line and not line.startswith("#"):
                    k, v = line.split("=", 1)
                    env[k.strip()] = v.strip()
            break

    # Normalize key names from ~/Desktop/.env format
    if "Datadog_API_KEY_Secret" in env and "datadog_key_credential" not in env:
        env["datadog_key_credential"] = env["Datadog_API_KEY_Secret"]
    if "Datadog_APP_KEY_Secret" in env and "datadog_app_key_credential" not in env:
        env["datadog_app_key_credential"] = env["Datadog_APP_KEY_Secret"]
    if "Zscaler_ID" in env and "ZSCALER_API_ADMIN" not in env:
        env["ZSCALER_API_ADMIN"] = env["Zscaler_ID"]
    if "Zscaler_PW" in env and "ZSCALER_API_PASSWORD" not in env:
        env["ZSCALER_API_PASSWORD"] = env["Zscaler_PW"]

    return env


env_vars: dict[str, str] = load_env()
DD_API_KEY = env_vars.get("datadog_key_credential", "")
DD_APP_KEY = env_vars.get("datadog_app_key_credential", "")


# ═══════════════════════════════════════════════════════════════════════════
# Google Sheets 수집
# ═══════════════════════════════════════════════════════════════════════════

def collect_sheets() -> dict[str, Any]:
    """Google Sheets API로 자산관리대장 + AI구독현황 수집"""
    import gspread

    print("\n[Sheets] Google Sheets 연결...")
    gc = gspread.oauth(
        scopes=[
            "https://www.googleapis.com/auth/spreadsheets.readonly",
            "https://www.googleapis.com/auth/drive.readonly",
        ]
    )

    result: dict[str, Any] = {
        "saas": [],
        "license": [],
        "ai": [],
        "asset_counts": {},
        "infra": {
            "servers": [],
            "databases": [],
            "sec_systems": [],
            "networks": [],
        },
    }

    # ── 자산관리대장 ──
    sp1 = gc.open_by_key(SHEET_IDS["자산관리대장"])
    print(f"  '{sp1.title}' 연결 OK")

    # SaaS 파싱
    ws_saas = sp1.worksheet("8.SaaS")
    raw = ws_saas.get_all_values()
    saas = []
    for row in raw[4:]:
        cells = [c.strip() for c in row]
        if not any(cells):
            continue
        auth_val = cells[7] if len(cells) > 7 else ""
        auth_lower = auth_val.lower()
        sso_linked = auth_lower in ("okta", "okta sso", "google")
        saas.append({
            "no": cells[1] if len(cells) > 1 else "",
            "category": cells[2] if len(cells) > 2 else "",
            "name": cells[3] if len(cells) > 3 else "",
            "description": cells[4] if len(cells) > 4 else "",
            "provider": cells[5] if len(cells) > 5 else "",
            "owner": cells[6] if len(cells) > 6 else "",
            "auth": auth_val,
            "sso": sso_linked,
            "contract_period": cells[9] if len(cells) > 9 else "",
            "note": cells[8] if len(cells) > 8 else "",
        })
    result["saas"] = saas
    print(f"    SaaS: {len(saas)}개")

    # 라이선스 현황
    ws_lic = sp1.worksheet("라이선스_현황")
    raw = ws_lic.get_all_values()
    license_list = []
    for row in raw[1:]:
        cells = [c.strip() for c in row]
        if not any(cells):
            continue
        license_list.append({
            "name": cells[0] if len(cells) > 0 else "",
            "description": cells[1] if len(cells) > 1 else "",
            "status": cells[2] if len(cells) > 2 else "",
            "admin": cells[3] if len(cells) > 3 else "",
            "department": cells[4] if len(cells) > 4 else "",
            "accounts": cells[5] if len(cells) > 5 else "",
            "active_accounts": cells[6] if len(cells) > 6 else "",
            "users": cells[7] if len(cells) > 7 else "",
            "cycle": cells[8] if len(cells) > 8 else "",
            "cost": cells[9] if len(cells) > 9 else "",
            "contract_period": cells[10] if len(cells) > 10 else "",
        })
    result["license"] = license_list
    print(f"    라이선스: {len(license_list)}개")

    # 자산 분류별 수량
    sheet_map = {
        "1.서버": "서버",
        "2.정보보호시스템": "정보보호시스템",
        "3.네트워크 장비": "네트워크 장비",
        "4. DBMS": "DBMS",
        "5.PC": "PC",
        "6.홈페이지": "홈페이지",
        "7. 개인정보": "개인정보",
        "8.SaaS": "SaaS",
    }
    asset_counts = {}
    for ws_item in sp1.worksheets():
        if ws_item.title in sheet_map:
            vals = ws_item.get_all_values()
            data_rows = len([r for r in vals[8:] if any(c.strip() for c in r)]) if len(vals) > 8 else 0
            asset_counts[sheet_map[ws_item.title]] = data_rows
    asset_counts["라이선스"] = len(license_list)
    result["asset_counts"] = asset_counts
    print(f"    자산 분류: {asset_counts}")

    # 서버 (EC2)
    raw_srv = sp1.worksheet("1.서버").get_all_values()
    servers = []
    for row in raw_srv[8:]:
        c = [x.strip() for x in row]
        if not c[1] and not c[2]:
            continue
        servers.append({
            "os_type": c[1], "instance_id": c[2], "name": c[3],
            "account": c[4], "os": c[5], "instance_type": c[6],
            "count": c[7], "region": c[8],
        })
    result["infra"]["servers"] = servers
    print(f"    서버 (EC2): {len(servers)}개")

    # DBMS
    raw_db = sp1.worksheet("4. DBMS").get_all_values()
    databases = []
    for row in raw_db[8:]:
        c = [x.strip() for x in row]
        if not c[1] and not c[2]:
            continue
        databases.append({
            "type": c[1], "name": c[2], "version": c[3],
            "engine": c[4], "count": c[6], "platform": c[7],
            "region": c[8], "account": c[10] if len(c) > 10 else "",
        })
    result["infra"]["databases"] = databases
    print(f"    DBMS: {len(databases)}개")

    # 정보보호시스템
    raw_sec = sp1.worksheet("2.정보보호시스템").get_all_values()
    sec_systems = []
    for row in raw_sec[8:]:
        c = [x.strip() for x in row]
        if not c[1] and not c[2]:
            continue
        sec_systems.append({
            "category": c[1], "name": c[2], "description": c[3],
            "type": c[4], "provider": c[5], "domain": c[6],
            "owner": c[7], "auth": c[8],
        })
    result["infra"]["sec_systems"] = sec_systems
    print(f"    정보보호시스템: {len(sec_systems)}개")

    # 네트워크 장비
    raw_net = sp1.worksheet("3.네트워크 장비").get_all_values()
    networks = []
    for row in raw_net[8:]:
        c = [x.strip() for x in row]
        if not c[1] and not c[2]:
            continue
        networks.append({
            "type": c[1], "resource_id": c[2], "name": c[3],
            "description": c[4], "region": c[5],
            "status": c[7], "account": c[9] if len(c) > 9 else "",
        })
    result["infra"]["networks"] = networks
    print(f"    네트워크 장비: {len(networks)}개")

    # ── AI 구독 현황 ──
    sp2 = gc.open_by_key(SHEET_IDS["AI구독현황"])
    print(f"  '{sp2.title}' 연결 OK")
    ws_ai = sp2.worksheets()[0]
    raw_ai = ws_ai.get_all_values()
    ai_list = []
    for row in raw_ai[3:]:
        cells = [c.strip() for c in row]
        if not any(cells):
            continue
        ai_list.append({
            "name": cells[0] if len(cells) > 0 else "",
            "description": cells[1] if len(cells) > 1 else "",
            "quantity": cells[2] if len(cells) > 2 else "",
            "department": cells[3] if len(cells) > 3 else "",
        })
    result["ai"] = ai_list
    print(f"    AI 구독: {len(ai_list)}개")

    return result


# ═══════════════════════════════════════════════════════════════════════════
# 비용 xlsx 파싱
# ═══════════════════════════════════════════════════════════════════════════

def parse_cost_xlsx(xlsx_path: Path) -> dict:
    """xlsx 파일에서 saas_cost 데이터 구조를 추출합니다."""
    if not openpyxl:
        print("  openpyxl 미설치, SaaS 비용 스킵")
        return {"summary": [], "details": {}, "by_user": [], "by_department": []}

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    cost_data: dict[str, Any] = {"summary": [], "details": {}, "by_user": [], "by_department": []}

    # ── 요약 시트 ──
    ws_sum = wb["요약"]
    section = "software"
    month_names = ["", "jan", "feb", "mar", "apr", "may", "jun",
                    "jul", "aug", "sep", "oct", "nov", "dec"]
    for row in ws_sum.iter_rows(min_row=3, values_only=True):
        cells = list(row)
        label = str(cells[0] or "").strip()
        if not label:
            continue
        if label in ("사용자별 요약", "사용자"):
            section = "user"
            continue
        if label in ("부서별 요약", "부서"):
            section = "department"
            continue
        if label in ("합계", "소프트웨어"):
            continue

        vals = {}
        for i in range(1, len(cells) - 1):
            v = cells[i] if isinstance(cells[i], (int, float)) else 0
            if i < len(month_names):
                vals[month_names[i]] = v
        total = cells[len(cells) - 2] if isinstance(cells[len(cells) - 2], (int, float)) else 0
        vals["total"] = total

        if section == "software":
            cost_data["summary"].append({"software": label, **vals})
        elif section == "user":
            cost_data["by_user"].append({"user": label, **vals})
        elif section == "department":
            cost_data["by_department"].append({"department": label, **vals})

    # ── 월별 상세 ──
    for sn in wb.sheetnames:
        if not sn.startswith("20"):
            continue
        ws_m = wb[sn]
        rows_m = []
        for row in ws_m.iter_rows(min_row=5, values_only=True):
            cells = list(row)
            if not cells[0]:
                continue
            sw = str(cells[0]).strip()
            if sw in ("합계",):
                continue
            date_val = cells[5] if len(cells) > 5 else ""
            if hasattr(date_val, "strftime"):
                date_val = date_val.strftime("%Y-%m-%d")
            rows_m.append({
                "software": sw,
                "user": str(cells[1] or ""),
                "shared": str(cells[2] or ""),
                "department": str(cells[3] or ""),
                "amount": cells[4] if isinstance(cells[4], (int, float)) else 0,
                "date": str(date_val or ""),
                "vendor": str(cells[6] or "") if len(cells) > 6 else "",
                "memo": str(cells[7] or "") if len(cells) > 7 else "",
                "status": str(cells[8] or "") if len(cells) > 8 else "",
            })
        cost_data["details"][sn] = rows_m

    # ── 중복/유사 항목 통합 ──
    merged: dict[str, dict] = {}
    for s in cost_data["summary"]:
        name = SW_MERGE.get(s["software"], s["software"])
        if name in merged:
            for k in list(s.keys()):
                if k != "software" and isinstance(s[k], (int, float)):
                    merged[name][k] = merged[name].get(k, 0) + s[k]
        else:
            merged[name] = {**s, "software": name}
    cost_data["summary"] = list(merged.values())

    for month_key in list(cost_data["details"].keys()):
        rows = cost_data["details"][month_key]
        seen: set[tuple] = set()
        deduped = []
        for r in rows:
            r["software"] = SW_MERGE.get(r["software"], r["software"])
            key = (r["software"], r.get("user", ""), r.get("amount", 0))
            if key in seen:
                continue
            seen.add(key)
            deduped.append(r)
        cost_data["details"][month_key] = deduped

    return cost_data


# ═══════════════════════════════════════════════════════════════════════════
# Datadog API
# ═══════════════════════════════════════════════════════════════════════════

def dd_get(path: str, params: str = "") -> dict[str, object]:
    url = f"https://api.datadoghq.com{path}"
    if params:
        url += f"?{params}"
    req = urllib.request.Request(
        url,
        headers={
            "DD-API-KEY": DD_API_KEY,
            "DD-APPLICATION-KEY": DD_APP_KEY,
            "Content-Type": "application/json",
        },
    )
    try:
        with urllib.request.urlopen(
            req, timeout=30
        ) as r:  # nosemgrep: dynamic-urllib-use-detected — trusted internal API URLs
            payload = json.loads(r.read().decode("utf-8"))
            return payload if isinstance(payload, dict) else {}
    except Exception as e:
        print(f"  DD API 오류 ({path}): {e}")
        return {}


def dd_post(path: str, body: object) -> dict[str, object]:
    url = f"https://api.datadoghq.com{path}"
    req = urllib.request.Request(
        url,
        json.dumps(body).encode(),
        method="POST",
        headers={
            "DD-API-KEY": DD_API_KEY,
            "DD-APPLICATION-KEY": DD_APP_KEY,
            "Content-Type": "application/json",
        },
    )
    try:
        with urllib.request.urlopen(
            req, timeout=30
        ) as r:  # nosemgrep: dynamic-urllib-use-detected — trusted internal API URLs
            payload = json.loads(r.read().decode("utf-8"))
            return payload if isinstance(payload, dict) else {}
    except Exception as e:
        print(f"  DD API 오류 ({path}): {e}")
        return {}


def collect_datadog():
    """Datadog 호스트 + 보안 시그널 수집"""
    if not DD_API_KEY:
        print("  Datadog API Key 없음, 스킵")
        return [], []

    print("\n[Datadog] 호스트...")
    raw = dd_get("/api/v1/hosts", "count=200")
    hosts = []
    host_list = raw.get("host_list", [])
    if not isinstance(host_list, list):
        host_list = []
    for h in host_list:
        if not isinstance(h, dict):
            continue
        tags = {}
        tags_by_source = h.get("tags_by_source", {})
        if not isinstance(tags_by_source, dict):
            tags_by_source = {}
        for src, tlist in tags_by_source.items():
            if not isinstance(tlist, list):
                continue
            for t in tlist:
                if ":" in t:
                    k, v = t.split(":", 1)
                    tags[k] = v
        hosts.append(
            {
                "name": h.get("name", ""),
                "instance_type": tags.get("instance-type", ""),
                "region": tags.get("region", ""),
                "cluster": tags.get(
                    "aws_eks_cluster-name", tags.get("eks_eks-cluster-name", "")
                ),
                "nodepool": tags.get("karpenter.sh/nodepool", ""),
                "env": tags.get("env", ""),
                "aws_alias": tags.get("aws_alias", ""),
                "aws_account": tags.get("aws_account", ""),
                "up": h.get("up", False),
                "agent_version": h.get("meta", {}).get("agent_version", ""),
            }
        )
    print(f"    {len(hosts)}개")

    print("  Datadog 보안 시그널 (14일)...")
    raw = dd_post(
        "/api/v2/security_monitoring/signals/search",
        {
            "filter": {
                "from": "now-14d",
                "to": "now",
                "query": "status:(high OR critical OR medium)",
            },
            "sort": "timestamp",
            "page": {"limit": 100},
        },
    )
    signals = []
    signal_data = raw.get("data", [])
    if not isinstance(signal_data, list):
        signal_data = []
    for s in signal_data:
        if not isinstance(s, dict):
            continue
        a = s.get("attributes", {})
        if not isinstance(a, dict):
            a = {}
        signals.append(
            {
                "id": s.get("id", "")[:30],
                "title": a.get("message", "")[:200],
                "severity": a.get("severity", ""),
                "status": a.get("status", ""),
                "timestamp": a.get("timestamp", ""),
                "tags": a.get("tags", [])[:5],
            }
        )
    print(f"    {len(signals)}건")
    return hosts, signals


# ═══════════════════════════════════════════════════════════════════════════
# AWS 수집
# ═══════════════════════════════════════════════════════════════════════════

def collect_aws():
    """Run AWS describe for EC2, RDS, EKS across configured profiles."""
    profiles = [p.strip() for p in env_vars.get("AWS_PROFILES", "").split(",") if p.strip()]
    if not profiles:
        print("\n[AWS] AWS_PROFILES 미설정, 스킵")
        return

    print(f"\n[AWS] 프로파일: {', '.join(profiles)}")
    commands = [
        ("aws ec2 describe-instances --profile {p} --output json", "Reservations", "aws-ec2-{p}.json"),
        ("aws rds describe-db-instances --profile {p} --output json", "DBInstances", "aws-rds-{p}.json"),
        ("aws eks list-clusters --profile {p} --output json", "clusters", "aws-eks-{p}.json"),
    ]
    for profile in profiles:
        for cmd_template, output_key, filename_template in commands:
            cmd = cmd_template.format(p=profile)
            filename = filename_template.format(p=profile)
            try:
                result = subprocess.run(
                    cmd.split(), capture_output=True, text=True, timeout=60
                )
                if result.returncode != 0:
                    print(f"    {filename}: 실패 ({result.stderr.strip()[:100]})")
                    continue
                data = json.loads(result.stdout)
                # EC2: flatten Reservations → instances
                if output_key == "Reservations":
                    instances = []
                    for r in data.get("Reservations", []):
                        for inst in r.get("Instances", []):
                            name = ""
                            for tag in inst.get("Tags", []):
                                if tag.get("Key") == "Name":
                                    name = tag.get("Value", "")
                            instances.append({
                                "InstanceId": inst.get("InstanceId", ""),
                                "Name": name,
                                "Type": inst.get("InstanceType", ""),
                                "State": inst.get("State", {}).get("Name", ""),
                                "PrivateIp": inst.get("PrivateIpAddress", ""),
                                "LaunchTime": str(inst.get("LaunchTime", "")),
                                "_profile": profile,
                            })
                    out_path = ASSETS_DIR / filename
                    out_path.write_text(json.dumps(instances, ensure_ascii=False, default=str))
                    print(f"    {filename}: {len(instances)}개 인스턴스")
                else:
                    items = data.get(output_key, [])
                    out_path = ASSETS_DIR / filename
                    if isinstance(items, list):
                        out_path.write_text(json.dumps(items, ensure_ascii=False, default=str))
                        print(f"    {filename}: {len(items)}개")
                    else:
                        out_path.write_text(json.dumps(data, ensure_ascii=False, default=str))
                        print(f"    {filename}: 저장 완료")
            except subprocess.TimeoutExpired:
                print(f"    {filename}: 타임아웃")
            except Exception as e:
                print(f"    {filename}: 오류 ({e})")


def load_aws_live_data():
    """Load AWS describe results from .claudesec-assets/aws-*.json files"""
    result = {"ec2": [], "rds": [], "elasticache": [], "s3": [], "eks": []}
    aws_profiles = [
        p.strip() for p in env_vars.get("AWS_PROFILES", "").split(",") if p.strip()
    ]
    if not aws_profiles:
        # Fallback: scan .claudesec-assets for aws-ec2-*.json files
        for f in sorted(ASSETS_DIR.glob("aws-ec2-*.json")):
            p = f.stem.replace("aws-ec2-", "")
            if p not in aws_profiles:
                aws_profiles.append(p)
    for profile in aws_profiles:
        for rtype in ["ec2", "rds", "rds-clusters", "elasticache", "s3", "eks"]:
            fpath = ASSETS_DIR / f"aws-{rtype}-{profile}.json"
            if not fpath.exists():
                continue
            try:
                data = json.loads(fpath.read_text())
                if isinstance(data, list):
                    for item in data:
                        item["_profile"] = profile
                    key = rtype.replace("-clusters", "")
                    if key not in result:
                        result[key] = []
                    result[key].extend(data)
                elif isinstance(data, dict) and "clusters" in data:
                    for c in data["clusters"]:
                        result["eks"].append({"name": c, "_profile": profile})
            except Exception:
                pass
    for k, v in result.items():
        if v:
            print(f"    AWS {k}: {len(v)}개")
    return result


def _is_karpenter_node(name: str) -> bool:
    """Karpenter 동적 노드 여부 판별 (ip-10-* 패턴)"""
    if not name:
        return False
    return bool(re.match(r"ip-10-\d+-\d+-\d+\.", name))


def _cross_verify_ec2(
    aws_ec2: list[dict[str, object]], sheet_servers: list[dict[str, object]]
) -> dict[str, Any]:
    """AWS EC2 실시간 vs 자산관리대장 서버 교차 검증 (Karpenter 노드 분리)"""
    # Karpenter 노드와 고정 서버 분리
    aws_static: dict[str, dict[str, object]] = {}
    aws_karpenter: list[dict[str, object]] = []
    for e in aws_ec2:
        iid = e.get("InstanceId", "")
        name = e.get("Name", "")
        if not isinstance(iid, str) or not iid:
            continue
        if not isinstance(name, str):
            name = ""
        if _is_karpenter_node(name):
            aws_karpenter.append(
                {
                    "id": iid,
                    "name": name,
                    "type": e.get("Type", ""),
                    "profile": e.get("_profile", ""),
                }
            )
        else:
            aws_static[iid] = e

    # 시트에서도 Karpenter/고정 분리
    sheet_static: dict[str, dict[str, object]] = {}
    sheet_karpenter: list[dict[str, object]] = []
    for s in sheet_servers:
        iid = s.get("instance_id", "")
        name = s.get("name", "")
        if not isinstance(iid, str) or not iid:
            continue
        if not isinstance(name, str):
            name = ""
        if _is_karpenter_node(name):
            sheet_karpenter.append(
                {
                    "id": iid,
                    "name": name,
                    "type": s.get("instance_type", ""),
                    "account": s.get("account", ""),
                }
            )
        else:
            sheet_static[iid] = s

    # 고정 서버 교차 검증
    aws_only: list[dict[str, object]] = []
    for iid, e in aws_static.items():
        if iid not in sheet_static:
            aws_only.append(
                {
                    "id": iid,
                    "name": e.get("Name", ""),
                    "type": e.get("Type", ""),
                    "profile": e.get("_profile", ""),
                    "status": "미등록",
                }
            )

    sheet_only: list[dict[str, object]] = []
    for iid, s in sheet_static.items():
        if iid not in aws_static:
            sheet_only.append(
                {
                    "id": iid,
                    "name": s.get("name", ""),
                    "type": s.get("instance_type", ""),
                    "account": s.get("account", ""),
                    "status": "종료/교체",
                }
            )

    matched = len(set(aws_static) & set(sheet_static))

    result = {
        "aws_total": len(aws_ec2),
        "aws_static": len(aws_static),
        "aws_karpenter": len(aws_karpenter),
        "sheet_total": len(sheet_servers),
        "sheet_static": len(sheet_static),
        "sheet_karpenter": len(sheet_karpenter),
        "matched": matched,
        "aws_only": aws_only,
        "sheet_only": sheet_only,
        "aws_only_count": len(aws_only),
        "sheet_only_count": len(sheet_only),
        "karpenter_nodes": aws_karpenter,
        "karpenter_count": len(aws_karpenter),
    }
    print(
        f"  교차검증: 고정 서버 AWS {len(aws_static)} vs 시트 {len(sheet_static)}, "
        f"일치 {matched}, 미등록 {len(aws_only)}, 종료 {len(sheet_only)}"
    )
    print(f"  Karpenter: AWS {len(aws_karpenter)}개 동적 노드 (교차검증 제외)")
    return result


# ═══════════════════════════════════════════════════════════════════════════
# Zscaler 수집
# ═══════════════════════════════════════════════════════════════════════════

def collect_zscaler() -> dict:
    """Zscaler ZIA API로 보안 포스처 수집"""
    api_key = env_vars.get("ZSCALER_API_KEY", "")
    admin = env_vars.get("ZSCALER_API_ADMIN", "")
    password = env_vars.get("ZSCALER_API_PASSWORD", "")
    base_url = env_vars.get("ZSCALER_BASE_URL", "")

    if not all([api_key, admin, password, base_url]):
        print("\n[Zscaler] 자격 증명 미설정, 스킵")
        return {}

    print("\n[Zscaler] ZIA API 연결...")

    # Set env vars and use the existing zscaler-api module
    os.environ["ZSCALER_API_KEY"] = api_key
    os.environ["ZSCALER_API_ADMIN"] = admin
    os.environ["ZSCALER_API_PASSWORD"] = password
    os.environ["ZSCALER_BASE_URL"] = base_url

    try:
        # Import from scanner/lib/zscaler-api.py (already in sys.path)
        import importlib
        zscaler_mod = importlib.import_module("zscaler-api")

        import requests as req_lib
        session = req_lib.Session()
        if not zscaler_mod._auth(session, base_url, api_key, admin, password):
            print("    인증 실패")
            return {"error": "auth_failed"}

        try:
            posture = zscaler_mod.collect_posture(base_url, session)
            posture["authenticated"] = True
            print(f"    수집 완료: {list(posture.keys())}")
            return posture
        finally:
            try:
                session.delete(f"{base_url}/api/v1/authenticatedSession", timeout=5)
            except Exception:
                pass
    except ImportError as e:
        print(f"    모듈 로드 실패: {e}")
        return {"error": str(e)}
    except Exception as e:
        print(f"    수집 실패: {e}")
        return {"error": str(e)}


# ═══════════════════════════════════════════════════════════════════════════
# Prowler 수집
# ═══════════════════════════════════════════════════════════════════════════

def collect_prowler():
    """Prowler 결과 요약"""
    print("\n[Prowler] 결과 분석...")
    pdir = ROOT / ".claudesec-prowler"
    if not pdir.exists():
        print("    디렉토리 없음, 스킵")
        return {}
    findings = []
    for f in pdir.glob("*.ocsf.json"):
        try:
            data = json.loads(f.read_text())
            findings.extend(data if isinstance(data, list) else [data])
        except json.JSONDecodeError:
            # Handle NDJSON or multi-object files
            decoder = json.JSONDecoder()
            raw = f.read_text().strip()
            pos = 0
            while pos < len(raw):
                try:
                    obj, end = decoder.raw_decode(raw, pos)
                    if isinstance(obj, list):
                        findings.extend(obj)
                    else:
                        findings.append(obj)
                    pos = end
                except json.JSONDecodeError:
                    pos += 1
        except Exception:
            pass

    if not findings:
        print("    결과 없음")
        return {}

    failures = [f for f in findings if f.get("status_code") == "FAIL"]

    def _is_critical_or_high(f):
        return (f.get("severity") or "").lower() in ("high", "critical")

    _PROWLER_NAME_FIX = {
        "core_minimize_containers_added_capabiliti": "core_minimize_containers_added_capabilities",
        "iam_aws_attached_policy_no_administrative_privil": "iam_aws_attached_policy_no_administrative_privileges",
    }
    _PROWLER_HUB_BLOCKLIST = {"iac-branch", "slack-web-hook"}

    def _prowler_hub_url(check_id: str) -> str:
        if not check_id:
            return ""
        name = re.sub(r"^prowler-[a-z]+-", "", check_id)
        name = re.sub(r"-iac-branch-\.[a-z0-9/]+$", "", name)
        name = re.sub(r"-\d{12}.*$", "", name)
        name = re.sub(r"-[0-9a-f]{5,}$", "", name)
        if any(bl in name for bl in _PROWLER_HUB_BLOCKLIST):
            return ""
        name = _PROWLER_NAME_FIX.get(name, name)
        return f"https://hub.prowler.com/check/{name}" if name else ""

    fail_sev = {}
    for f in failures:
        s = f.get("severity", "unknown")
        fail_sev[s] = fail_sev.get(s, 0) + 1

    samples = []
    seen_msgs = set()
    for f in failures:
        if _is_critical_or_high(f):
            msg = f.get("message", f.get("status_detail", ""))[:150]
            if msg in seen_msgs:
                continue
            seen_msgs.add(msg)
            um = f.get("unmapped", {})
            check_id = (
                f.get("finding_info", {}).get("uid", "")
                if isinstance(f.get("finding_info"), dict)
                else ""
            )[:60]
            samples.append(
                {
                    "message": msg,
                    "severity": f.get("severity", "High"),
                    "provider": um.get("provider", "") if isinstance(um, dict) else "",
                    "check": check_id,
                    "hub_url": _prowler_hub_url(check_id),
                }
            )
    samples.sort(key=lambda x: (0 if x["severity"] == "Critical" else 1, x["message"]))

    # Medium samples
    medium_samples = []
    critical_msgs = {s["message"] for s in samples}
    for f in failures:
        msg = f.get("message", f.get("status_detail", ""))[:150]
        if (
            f.get("severity") in ("Medium", "High")
            and msg not in critical_msgs
            and len(medium_samples) < 50
        ):
            um = f.get("unmapped", {})
            mid = (
                f.get("finding_info", {}).get("uid", "")
                if isinstance(f.get("finding_info"), dict)
                else ""
            )[:60]
            medium_samples.append(
                {
                    "message": msg,
                    "severity": f.get("severity", "Medium"),
                    "provider": um.get("provider", "") if isinstance(um, dict) else "",
                    "check": mid,
                    "hub_url": _prowler_hub_url(mid),
                }
            )

    # Group by provider
    by_provider = {}
    for f in failures:
        um = f.get("unmapped", {})
        prov = (
            um.get("provider", "") if isinstance(um, dict) else ""
        ).lower() or "unknown"
        if prov not in by_provider:
            by_provider[prov] = {"total": 0, "fail": 0, "pass": 0, "critical_fails": []}
        by_provider[prov]["fail"] += 1

    for f in findings:
        um = f.get("unmapped", {})
        prov = (
            um.get("provider", "") if isinstance(um, dict) else ""
        ).lower() or "unknown"
        if prov not in by_provider:
            by_provider[prov] = {"total": 0, "fail": 0, "pass": 0, "critical_fails": []}
        by_provider[prov]["total"] += 1
        if f.get("status_code") != "FAIL":
            by_provider[prov]["pass"] += 1

    for f in failures:
        if _is_critical_or_high(f):
            um = f.get("unmapped", {})
            prov = (
                um.get("provider", "") if isinstance(um, dict) else ""
            ).lower() or "unknown"
            if prov in by_provider and len(by_provider[prov]["critical_fails"]) < 10:
                cid = (
                    f.get("finding_info", {}).get("uid", "")[:60]
                    if isinstance(f.get("finding_info"), dict)
                    else ""
                )
                by_provider[prov]["critical_fails"].append(
                    {
                        "message": f.get("message", f.get("status_detail", ""))[:150],
                        "severity": f.get("severity", "High"),
                        "check": cid,
                        "hub_url": _prowler_hub_url(cid),
                    }
                )

    # Normalize provider keys
    PROVIDER_MERGE = {"k8s": "kubernetes", "iac": "IaC"}
    normalized_by_provider = {}
    for prov, stats in by_provider.items():
        canonical = PROVIDER_MERGE.get(prov, prov)
        if canonical not in normalized_by_provider:
            normalized_by_provider[canonical] = {"total": 0, "fail": 0, "pass": 0, "critical_fails": []}
        normalized_by_provider[canonical]["total"] += stats["total"]
        normalized_by_provider[canonical]["fail"] += stats["fail"]
        normalized_by_provider[canonical]["pass"] += stats["pass"]
        normalized_by_provider[canonical]["critical_fails"].extend(stats["critical_fails"])
    by_provider = normalized_by_provider

    result = {
        "total": len(findings),
        "pass": len(findings) - len(failures),
        "fail": len(failures),
        "fail_by_severity": fail_sev,
        "critical_fails_sample": samples[:30],
        "medium_fails_sample": medium_samples,
        "by_provider": by_provider,
    }
    print(f"    총 {len(findings)}건 (FAIL: {len(failures)})")
    return result


# ═══════════════════════════════════════════════════════════════════════════
# 스캔 히스토리
# ═══════════════════════════════════════════════════════════════════════════

def collect_scan_history():
    """Collect scan history from .claudesec-history/"""
    hist_dir = ROOT / ".claudesec-history"
    if not hist_dir.exists():
        return []
    history = []
    for f in sorted(hist_dir.glob("scan-*.json")):
        try:
            data = json.loads(f.read_text())
            history.append(data)
        except Exception:
            pass
    history.sort(key=lambda x: x.get("timestamp", ""))
    return history[-30:]


# ═══════════════════════════════════════════════════════════════════════════
# 대시보드 데이터 조합
# ═══════════════════════════════════════════════════════════════════════════

def build_dashboard_data(
    sheets_data: dict[str, Any] | None,
    cost_data: dict[str, Any] | None,
    scan_report: dict | None = None,
    dd_hosts: list | None = None,
    dd_signals: list | None = None,
    prowler: dict | None = None,
    scan_history: list | None = None,
    aws_live: dict | None = None,
    cross_verify: dict | None = None,
    zscaler_data: dict | None = None,
) -> dict:
    """dashboard-data.json 구조를 조합합니다."""
    # 기존 데이터 로드
    existing_path = ASSETS_DIR / "dashboard-data.json"
    data: dict[str, Any] = {}
    if existing_path.exists():
        try:
            data = json.loads(existing_path.read_text())
        except (json.JSONDecodeError, KeyError):
            pass

    # 기본 구조 보장
    data.setdefault("generated_at", NOW)
    data.setdefault("timestamps", {})
    data.setdefault("datadog", {"hosts": [], "signals": []})
    data.setdefault("prowler", {})
    data.setdefault("claudesec", {})
    data.setdefault("scan_history", [])
    data.setdefault("saas", [])
    data.setdefault("license", [])
    data.setdefault("ai", [])
    data.setdefault("asset_counts", {})
    data.setdefault("saas_cost", {})
    data.setdefault("infra", {})
    data.setdefault("aws_live", {})
    data.setdefault("cross_verify", {})
    data.setdefault("zscaler", {})
    data["generated_at"] = NOW

    # Sheets 데이터 머지
    if sheets_data:
        data["saas"] = sheets_data["saas"]
        data["license"] = sheets_data["license"]
        data["ai"] = sheets_data["ai"]
        data["asset_counts"] = sheets_data["asset_counts"]
        data["infra"] = sheets_data["infra"]
        data["timestamps"]["google_sheets"] = NOW

    # 비용 데이터 머지
    if cost_data:
        data["saas_cost"] = cost_data
        data["timestamps"]["saas_cost_xlsx"] = NOW

    # 스캔 리포트 머지
    if scan_report:
        data["claudesec"] = scan_report

    # Datadog 머지
    if dd_hosts is not None:
        data["datadog"]["hosts"] = dd_hosts
        data["timestamps"]["datadog_hosts"] = NOW
    if dd_signals is not None:
        data["datadog"]["signals"] = dd_signals
        data["timestamps"]["datadog_signals"] = NOW

    # Prowler 머지
    if prowler:
        data["prowler"] = prowler
        data["timestamps"]["prowler"] = NOW

    # 스캔 히스토리 머지
    if scan_history is not None:
        data["scan_history"] = scan_history

    # AWS 실시간 머지
    if aws_live:
        data["aws_live"] = aws_live
        data["timestamps"]["aws_live"] = NOW

    # 교차 검증 머지
    if cross_verify:
        data["cross_verify"] = cross_verify
        data["timestamps"]["cross_verify"] = NOW

    # Zscaler 머지
    if zscaler_data:
        data["zscaler"] = zscaler_data
        data["timestamps"]["zscaler"] = NOW

    return data


# ═══════════════════════════════════════════════════════════════════════════
# HTML 주입
# ═══════════════════════════════════════════════════════════════════════════

def inject_html(dashboard_data: dict):
    """대시보드 HTML에 데이터를 주입하여 live 파일 생성"""
    tmpl_path = ROOT / "claudesec-asset-dashboard.html"
    html = tmpl_path.read_text()
    json_str = json.dumps(dashboard_data, ensure_ascii=False, default=str)
    html = html.replace("__DASHBOARD_DATA__", json_str)
    nonce = generate_nonce()
    html = inject_csp_nonce(html, nonce)
    out_path = ROOT / "claudesec-asset-dashboard-live.html"
    out_path.write_text(html)
    print(f"  HTML 생성: {out_path}")


# ═══════════════════════════════════════════════════════════════════════════
# Main
# ═══════════════════════════════════════════════════════════════════════════

def main():
    import argparse
    parser = argparse.ArgumentParser(description="ClaudeSec 통합 대시보드 데이터 동기화")
    parser.add_argument("--xlsx", type=str, default=None, help="xlsx 파일 경로")
    parser.add_argument("--cost-only", action="store_true", help="비용 데이터만 동기화")
    parser.add_argument("--sheets-only", action="store_true", help="Sheets 데이터만 동기화")
    parser.add_argument("--inject-html", action="store_true", help="HTML 대시보드에 데이터 주입")
    parser.add_argument("--skip-aws", action="store_true", help="AWS 수집 스킵")
    parser.add_argument("--skip-datadog", action="store_true", help="Datadog 수집 스킵")
    parser.add_argument("--skip-zscaler", action="store_true", help="Zscaler 수집 스킵")
    args = parser.parse_args()

    print("━━━ ClaudeSec 통합 대시보드 데이터 동기화 ━━━")
    print(f"시각: {NOW}")

    # Sheet IDs를 환경변수에서 로드
    global SHEET_IDS
    SHEET_IDS = {
        "자산관리대장": env_vars.get("ASSET_SHEET_ID", ""),
        "AI구독현황": env_vars.get("AI_SHEET_ID", ""),
    }
    if not SHEET_IDS["자산관리대장"]:
        print("  ⚠ ASSET_SHEET_ID 환경변수 없음 — .env에 설정 필요")

    sheets_data = None
    cost_data = None
    dd_hosts = None
    dd_signals = None
    prowler = None
    scan_history = None
    aws_live = None
    cross_verify = None
    zscaler_data = None

    # ── Sheets 수집 ──
    if not args.cost_only:
        try:
            sheets_data = collect_sheets()
        except Exception as e:
            print(f"\n[Sheets] 수집 실패: {e}")
            print("  gspread 인증이 필요합니다. 아래 명령어로 인증하세요:")
            print("  python3 -c \"import gspread; gspread.oauth()\"")

    # ── 비용 xlsx 수집 ──
    if not args.sheets_only:
        xlsx_path = Path(args.xlsx) if args.xlsx else Path(env_vars.get("COST_XLSX_PATH", str(DEFAULT_XLSX)))
        if xlsx_path.exists():
            print(f"\n[Cost] 소스: {xlsx_path.name}")
            print(f"  수정일: {datetime.fromtimestamp(xlsx_path.stat().st_mtime, tz=KST).strftime('%Y-%m-%d %H:%M KST')}")
            cost_data = parse_cost_xlsx(xlsx_path)
            total_cost = sum(s.get("total", 0) for s in cost_data.get("summary", []))
            print(f"    소프트웨어: {len(cost_data['summary'])}종")
            print(f"    월별 상세:  {sum(len(v) for v in cost_data['details'].values())}건")
            print(f"    총 비용:    \\{total_cost:,.0f}")
        else:
            print(f"\n[Cost] 파일 없음: {xlsx_path}")

    # ── Datadog 수집 ──
    if not args.cost_only and not args.sheets_only and not args.skip_datadog:
        try:
            dd_hosts, dd_signals = collect_datadog()
        except Exception as e:
            print(f"\n[Datadog] 수집 실패: {e}")

    # ── AWS 수집 ──
    if not args.cost_only and not args.sheets_only and not args.skip_aws:
        try:
            collect_aws()
        except Exception as e:
            print(f"\n[AWS] 수집 실패: {e}")

        try:
            print("\n[AWS] 캐시 데이터 로드...")
            aws_live = load_aws_live_data()
        except Exception as e:
            print(f"\n[AWS] 로드 실패: {e}")

        # 교차 검증
        if aws_live and aws_live.get("ec2") and sheets_data and sheets_data.get("infra", {}).get("servers"):
            try:
                print("\n[교차검증] EC2 vs 자산관리대장...")
                cross_verify = _cross_verify_ec2(
                    aws_live["ec2"], sheets_data["infra"]["servers"]
                )
            except Exception as e:
                print(f"\n[교차검증] 실패: {e}")

    # ── Zscaler 수집 ──
    if not args.cost_only and not args.sheets_only and not args.skip_zscaler:
        try:
            zscaler_data = collect_zscaler()
        except Exception as e:
            print(f"\n[Zscaler] 수집 실패: {e}")

    # ── Prowler 수집 ──
    if not args.cost_only and not args.sheets_only:
        try:
            prowler = collect_prowler()
        except Exception as e:
            print(f"\n[Prowler] 수집 실패: {e}")

    # ── 스캔 히스토리 ──
    if not args.cost_only and not args.sheets_only:
        try:
            scan_history = collect_scan_history()
            if scan_history:
                print(f"\n[히스토리] {len(scan_history)}건")
        except Exception as e:
            print(f"\n[히스토리] 수집 실패: {e}")

    # 데이터가 하나도 없으면 종료
    has_data = any([
        sheets_data, cost_data, dd_hosts, dd_signals,
        prowler, aws_live, zscaler_data, scan_history,
    ])
    if not has_data:
        print("\n수집된 데이터 없음. 종료.")
        sys.exit(1)

    # scan-report.json 로드
    scan_report = None
    scan_path = ROOT / "scan-report.json"
    if scan_path.exists():
        try:
            scan_report = json.loads(scan_path.read_text())
        except Exception:
            pass

    # 대시보드 데이터 조합
    dashboard_data = build_dashboard_data(
        sheets_data=sheets_data,
        cost_data=cost_data,
        scan_report=scan_report,
        dd_hosts=dd_hosts,
        dd_signals=dd_signals,
        prowler=prowler,
        scan_history=scan_history,
        aws_live=aws_live,
        cross_verify=cross_verify,
        zscaler_data=zscaler_data,
    )

    # JSON 저장
    json_path = ASSETS_DIR / "dashboard-data.json"
    json_path.write_text(json.dumps(dashboard_data, ensure_ascii=False, indent=2, default=str))
    print(f"\n  저장: {json_path}")

    # HTML 주입
    if args.inject_html:
        inject_html(dashboard_data)

    # 요약
    print("\n━━━ 동기화 결과 ━━━")
    ts = dashboard_data.get("timestamps", {})
    if sheets_data:
        print(f"  SaaS:        {len(sheets_data['saas'])}개")
        print(f"  라이선스:     {len(sheets_data['license'])}개")
        print(f"  AI 구독:      {len(sheets_data['ai'])}개")
        print(f"  자산 분류:    {sheets_data['asset_counts']}")
        infra = sheets_data["infra"]
        print(f"  인프라:       서버 {len(infra['servers'])}, DB {len(infra['databases'])}, 보안 {len(infra['sec_systems'])}, 네트워크 {len(infra['networks'])}")
    if cost_data:
        print(f"  SaaS 비용:    {len(cost_data['summary'])}종, \\{sum(s.get('total', 0) for s in cost_data['summary']):,.0f}")
    if dd_hosts:
        print(f"  Datadog:      호스트 {len(dd_hosts)}개, 시그널 {len(dd_signals or [])}건")
    if aws_live:
        ec2_count = len(aws_live.get("ec2", []))
        rds_count = len(aws_live.get("rds", []))
        eks_count = len(aws_live.get("eks", []))
        print(f"  AWS:          EC2 {ec2_count}, RDS {rds_count}, EKS {eks_count}")
    if cross_verify:
        print(f"  교차검증:     일치 {cross_verify.get('matched', 0)}, 미등록 {cross_verify.get('aws_only_count', 0)}, 종료 {cross_verify.get('sheet_only_count', 0)}")
    if zscaler_data and not zscaler_data.get("error"):
        print(f"  Zscaler:      인증 {'OK' if zscaler_data.get('authenticated') else 'FAIL'}")
    if prowler:
        print(f"  Prowler:      총 {prowler.get('total', 0)}건 (FAIL: {prowler.get('fail', 0)})")
    if scan_history:
        print(f"  히스토리:     {len(scan_history)}건")

    print(f"\n  타임스탬프: {json.dumps(ts, ensure_ascii=False)}")
    print()


if __name__ == "__main__":
    main()
